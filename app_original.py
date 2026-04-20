from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, make_response
from functools import wraps
import sqlite3, hashlib, os, shutil, json, io
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cartridge-secret-key-2026-change-me")

DB_PATH = os.path.join(os.path.dirname(__file__), "data.db")
BACKUP_DIR = os.path.join(os.path.dirname(__file__), "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)

MONTHS = ["Январь","Февраль","Март","Апрель","Май","Июнь",
          "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

# ── Windows 11 compatibility check ────────────────────────────────────────────
def check_win11_compat(cpu, ram_str):
    import re
    cpu = (cpu or "").strip()
    ram_str = (ram_str or "").strip()
    if not cpu:
        return "unknown", "Нет данных о процессоре"
    cpu_u = cpu.upper()

    # Parse RAM (МБ или ГБ)
    ram_mb = 0
    rm = re.search(r'(\d+)', ram_str)
    if rm:
        val = int(rm.group(1))
        if "GB" in ram_str.upper() or "ГБ" in ram_str.upper():
            ram_mb = val * 1024
        elif val < 512:
            ram_mb = val * 1024
        else:
            ram_mb = val
    ram_ok = (ram_mb == 0) or (ram_mb >= 4096)

    cpu_ok = None
    reason = ""

    if "CORE(TM)2" in cpu_u or "CORE 2" in cpu_u:
        cpu_ok, reason = False, "Intel Core 2 — несовместим"
    elif "CELERON" in cpu_u:
        cm = re.search(r'G(\d{4})', cpu_u)
        if cm and int(cm.group(1)) >= 5900:
            cpu_ok, reason = True, f"Intel Celeron G{cm.group(1)}"
        else:
            cpu_ok, reason = False, "Intel Celeron — не в списке Win11"
    elif "PENTIUM" in cpu_u:
        pm = re.search(r'G(\d{4})', cpu_u)
        if pm:
            gnum = int(pm.group(1))
            if gnum >= 7000:
                cpu_ok, reason = True, f"Pentium Gold G{gnum} (12-е пок.)"
            else:
                cpu_ok, reason = False, f"Pentium Gold G{gnum} — вне списка Win11"
        else:
            cpu_ok, reason = False, "Intel Pentium — несовместим"
    elif "INTEL" in cpu_u:
        if "11TH GEN" in cpu_u:
            cpu_ok, reason = True, "Intel 11-е поколение"
        elif "12TH GEN" in cpu_u:
            cpu_ok, reason = True, "Intel 12-е поколение"
        elif "13TH GEN" in cpu_u:
            cpu_ok, reason = True, "Intel 13-е поколение"
        else:
            im = re.search(r'I[3579]-(\d{4,5})', cpu_u)
            if im:
                num = im.group(1)
                first2 = int(num[:2])
                gen = first2 if 8 <= first2 <= 14 else int(num[0])
                if gen >= 8:
                    cpu_ok, reason = True, f"Intel Core {gen}-е поколение"
                else:
                    cpu_ok, reason = False, f"Intel Core {gen}-е поколение — устарел"
            else:
                cpu_ok, reason = None, "Intel: не удалось определить поколение"
    elif "AMD" in cpu_u:
        if "RYZEN" in cpu_u:
            am = re.search(r'(\d{4})', cpu_u)
            if am and int(am.group(1)[0]) >= 3:
                cpu_ok, reason = True, f"AMD Ryzen серия {am.group(1)}"
            elif am:
                cpu_ok, reason = False, f"AMD Ryzen серия {am.group(1)} — устарел"
            else:
                cpu_ok, reason = True, "AMD Ryzen — совместим"
        elif "ATHLON" in cpu_u:
            cpu_ok, reason = False, "AMD Athlon — не в списке Win11"
        elif re.search(r'\bE[12]-', cpu_u) or "FX-" in cpu_u:
            cpu_ok, reason = False, "Старый AMD — несовместим"
        else:
            cpu_ok, reason = None, "AMD: не удалось определить"
    else:
        cpu_ok, reason = None, "Неизвестный процессор"

    if cpu_ok is False:
        return "incompatible", reason
    elif cpu_ok is True:
        if not ram_ok and ram_mb > 0:
            return "needs_upgrade", f"CPU OK ({reason}), RAM недостаточно ({ram_str})"
        return "compatible", reason
    else:
        return "unknown", reason

# ── DB helpers ────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS printers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            floor TEXT NOT NULL,
            dept TEXT NOT NULL,
            model TEXT NOT NULL,
            cartridge TEXT NOT NULL,
            monthly_rate REAL DEFAULT 0.5,
            price INTEGER DEFAULT 600,
            active INTEGER DEFAULT 1,
            note TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS toners (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            monthly_rate REAL DEFAULT 0.3,
            price INTEGER DEFAULT 1300,
            active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            printer_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            qty INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            updated_by TEXT DEFAULT '',
            UNIQUE(printer_id, year, month)
        );
        CREATE TABLE IF NOT EXISTS toner_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            toner_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            qty INTEGER DEFAULT 0,
            updated_by TEXT DEFAULT '',
            UNIQUE(toner_id, year, month)
        );
        CREATE TABLE IF NOT EXISTS stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cartridge TEXT NOT NULL UNIQUE,
            qty INTEGER DEFAULT 0,
            min_qty INTEGER DEFAULT 2,
            price INTEGER DEFAULT 0,
            note TEXT DEFAULT '',
            updated_at TEXT,
            updated_by TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS stock_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cartridge TEXT NOT NULL,
            change_qty INTEGER NOT NULL,
            reason TEXT DEFAULT '',
            created_at TEXT,
            created_by TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS equipment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            floor TEXT NOT NULL,
            dept TEXT NOT NULL,
            responsible TEXT DEFAULT '',
            pc_name TEXT DEFAULT '',
            pc_inv TEXT DEFAULT '',
            pc_brand TEXT DEFAULT '',
            pc_model TEXT DEFAULT '',
            pc_serial TEXT DEFAULT '',
            pc_os TEXT DEFAULT '',
            pc_cpu TEXT DEFAULT '',
            pc_ram TEXT DEFAULT '',
            pc_hdd TEXT DEFAULT '',
            pc_ip TEXT DEFAULT '',
            pc_mac TEXT DEFAULT '',
            monitor TEXT DEFAULT '',
            monitor_inv TEXT DEFAULT '',
            monitor_serial TEXT DEFAULT '',
            keyboard TEXT DEFAULT '',
            mouse TEXT DEFAULT '',
            ups TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            other_devices TEXT DEFAULT '',
            purchase_date TEXT DEFAULT '',
            warranty_until TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            note TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT,
            updated_by TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS equipment_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            old_value TEXT DEFAULT '',
            new_value TEXT DEFAULT '',
            changed_at TEXT,
            changed_by TEXT DEFAULT ''
        );
        """)
        # Default admin
        pw = hashlib.sha256("admin123".encode()).hexdigest()
        try:
            db.execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",("admin",pw,"admin"))
        except: pass
        # Seed printers from xlsx data
        if db.execute("SELECT COUNT(*) FROM printers").fetchone()[0] == 0:
            seed_printers(db)
        if db.execute("SELECT COUNT(*) FROM toners").fetchone()[0] == 0:
            seed_toners(db)

def seed_printers(db):
    data = [
        ("1 этаж","Отдел кадров","KYOCERA 1125","TK-1110",1,450),
        ("1 этаж","Отдел кадров","BROTHER L2365DWR","TN2375T",0.3,1000),
        ("1 этаж","Снабжение","RICOH","SP230",0.3,900),
        ("1 этаж","Метрология","Pantum 6550NW","211",0.3,900),
        ("2 этаж","Юрист","Brother 1202","TN-1095",0.5,450),
        ("2 этаж","Договорной отдел","Pantum 6552","212",2,900),
        ("2 этаж","Договорной отдел","Pantum 6500","212",0.5,900),
        ("2 этаж","Охрана труда","HP1020","12A",0.3,500),
        ("2 этаж","Охрана труда","Pantum 2000","211",0.3,900),
        ("2 этаж","ФЭО","Pantum 6550","211",0.5,900),
        ("2 этаж","ФЭО","HP1020","12A",0.2,500),
        ("2 этаж","ФЭО","HP1020","12A",0.5,500),
        ("2 этаж","ФЭО","KYOCERA 2235","TK-1150",1,600),
        ("2 этаж","Склад","Pantum","212",1,900),
        ("3 этаж","Директор по экономике","Pantum 6552","212",0.3,900),
        ("3 этаж","Приемная","KYOCERA 2135","TK-1200",1,600),
        ("3 этаж","Технический директор","HP MFP175","CSP-W1106",0.2,800),
        ("3 этаж","Генеральный директор","HP132","19A/18A",0.2,1000),
        ("4 этаж","Юрист","HP1020","12A",0.3,500),
        ("4 этаж","Военпреды","Brother","TN-1075",0.2,600),
        ("4 этаж","Нач. отдела перспективных разработок","Canon 3228","EP-27",0.2,800),
        ("4 этаж","Канцелярия","KYOCERA 2040","TK-1200",0.5,600),
        ("5 этаж","ОВТ и ПО","OKI 5450","5200-YMKC",0.2,2500),
        ("5 этаж","ОВТ и ПО","HP 179","Y,K,M,C",0.2,2500),
        ("5 этаж","Нач. оптического цеха","HP1020","12A",0.2,600),
        ("5 этаж","Нач. отдела","HP1020","12А",0.2,600),
        ("5 этаж","Диспетчер","Brother 1202","TN-1095",0.3,450),
        ("5 этаж","Ст. диспетчер","HP1020","12A",0.3,600),
        ("6 этаж","ОТИЗ","Pantum","212",0.5,600),
        ("6 этаж","БКК","Canon 3228","EP-27",0.2,800),
        ("6 этаж","Нач. отдела КК","HP1020","12A",0.2,600),
        ("6 этаж","Бухгалтерия","Pantum 6550","211",1,900),
        ("6 этаж","Бухгалтерия","Pantum 6550","211",1,900),
        ("6 этаж","Бухгалтерия","HP1020","12A",1,600),
        ("6 этаж","Бухгалтерия","HP1020","12A",0.5,600),
        ("6 этаж","Бухгалтерия","KYOCERA 2035","TK-1200",1,600),
        ("7 этаж","Гл. оптик","HP1020","12A",0.2,600),
        ("7 этаж","К-Т О","HP5000","29x",0.3,2500),
        ("7 этаж","Диспетчер","KYOCERA 1025","TK-1110",0.5,450),
        ("7 этаж","Нач. сборочного цеха","KYOCERA 1020","TK-1110",0.2,450),
        ("7 этаж","Диспетчер","KYOCERA 1025","TK-1110",0.5,450),
        ("7 этаж","Диспетчер (склад)","Pantum 6552","212",0.5,900),
        ("7 этаж","Директор по производству","Pantum P2500W","211",0.2,900),
        ("7 этаж","Инженер-технолог","Brother","NT-1000",0.2,0),
        ("8 этаж","БТД","KYOCERA 181","435",0.5,1500),
        ("8 этаж","Инженер-нормировщик","HP1020","12A",0.5,600),
        ("8 этаж","Зам. ген. директора по подготовке кадров","XEROX 3020","106R02773",0.2,800),
    ]
    db.executemany("INSERT INTO printers (floor,dept,model,cartridge,monthly_rate,price) VALUES (?,?,?,?,?,?)", data)

def seed_toners(db):
    data = [
        ("Тонер KYOCERA (A) 20хх",0.3,1300),
        ("Тонер KYOCERA (B) 10хх",0.3,1300),
        ("Тонер Brother",0.2,1300),
        ("Тонер SAMSUNG",0.3,1300),
        ("Тонер HP",0.3,1300),
    ]
    db.executemany("INSERT INTO toners (name,monthly_rate,price) VALUES (?,?,?)", data)

# ── Auth helpers ──────────────────────────────────────────────────────────────
def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            return jsonify({"error": "Недостаточно прав"}), 403
        return f(*args, **kwargs)
    return decorated

# ── Routes: Auth ──────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if "user_id" not in session:
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))

@app.route("/login", methods=["GET","POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        with get_db() as db:
            user = db.execute("SELECT * FROM users WHERE username=? AND password=?",
                              (username, hash_pw(password))).fetchone()
        if user:
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))
        else:
            error = "Неверный логин или пароль"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Routes: Pages ─────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")

@app.route("/records")
@login_required
def records_page():
    return render_template("records.html", months=MONTHS)

@app.route("/toners")
@login_required
def toners_page():
    return render_template("toners.html", months=MONTHS)

@app.route("/printers")
@login_required
def printers_page():
    return render_template("printers.html")

@app.route("/analytics")
@login_required
def analytics_page():
    return render_template("analytics.html", months=MONTHS)

@app.route("/admin")
@login_required
def admin_page():
    if session.get("role") != "admin":
        return redirect(url_for("dashboard"))
    return render_template("admin.html")

@app.route("/stock")
@login_required
def stock_page():
    return render_template("stock.html")

@app.route("/equipment")
@login_required
def equipment_page():
    resp = make_response(render_template("equipment.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp

# ── API: Dashboard stats ──────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def api_stats():
    year = request.args.get("year", datetime.now().year, type=int)
    with get_db() as db:
        printers = db.execute("SELECT * FROM printers WHERE active=1").fetchall()
        toners = db.execute("SELECT * FROM toners WHERE active=1").fetchall()

        by_month = []
        for m in range(12):
            qty = db.execute("""SELECT COALESCE(SUM(r.qty),0) FROM records r
                JOIN printers p ON r.printer_id=p.id
                WHERE r.year=? AND r.month=? AND p.active=1""", (year, m)).fetchone()[0]
            cost = db.execute("""SELECT COALESCE(SUM(r.qty*p.price),0) FROM records r
                JOIN printers p ON r.printer_id=p.id
                WHERE r.year=? AND r.month=? AND p.active=1""", (year, m)).fetchone()[0]
            t_qty = db.execute("""SELECT COALESCE(SUM(r.qty),0) FROM toner_records r
                JOIN toners t ON r.toner_id=t.id
                WHERE r.year=? AND r.month=? AND t.active=1""", (year, m)).fetchone()[0]
            t_cost = db.execute("""SELECT COALESCE(SUM(r.qty*t.price),0) FROM toner_records r
                JOIN toners t ON r.toner_id=t.id
                WHERE r.year=? AND r.month=? AND t.active=1""", (year, m)).fetchone()[0]
            by_month.append({"qty": qty+t_qty, "cost": cost+t_cost})

        total_spent = sum(m["cost"] for m in by_month)
        total_items = sum(m["qty"] for m in by_month)

        expected_annual = sum(p["monthly_rate"]*p["price"]*12 for p in printers) + \
                          sum(t["monthly_rate"]*t["price"]*12 for t in toners)

        by_floor = {}
        for p in printers:
            fl = p["floor"]
            row = db.execute("""SELECT COALESCE(SUM(qty),0) as qty FROM records
                WHERE printer_id=? AND year=?""", (p["id"], year)).fetchone()
            cost = row["qty"] * p["price"]
            if fl not in by_floor:
                by_floor[fl] = {"qty": 0, "cost": 0}
            by_floor[fl]["qty"] += row["qty"]
            by_floor[fl]["cost"] += cost

        # Расход по помещениям (dept)
        dept_rows = db.execute("""
            SELECT p.floor, p.dept,
                   COUNT(p.id) as printer_count,
                   COALESCE(SUM(r.qty),0) as qty,
                   COALESCE(SUM(r.qty*p.price),0) as cost
            FROM printers p
            LEFT JOIN records r ON r.printer_id=p.id AND r.year=?
            WHERE p.active=1
            GROUP BY p.floor, p.dept
            ORDER BY cost DESC, qty DESC
        """, (year,)).fetchall()
        by_dept = [dict(r) for r in dept_rows]

        # ТОП принтеров по расходу
        top_printers = db.execute("""
            SELECT p.id, p.model, p.cartridge, p.dept, p.floor, p.price,
                   COALESCE(SUM(r.qty),0) as qty,
                   COALESCE(SUM(r.qty*p.price),0) as cost
            FROM printers p
            LEFT JOIN records r ON r.printer_id=p.id AND r.year=?
            WHERE p.active=1
            GROUP BY p.id
            ORDER BY qty DESC
            LIMIT 7
        """, (year,)).fetchall()

        # Прогноз: среднее по заполненным месяцам * 12
        cur_month_idx = datetime.now().month - 1  # 0-indexed
        if year == datetime.now().year:
            done_months = [by_month[i]["cost"] for i in range(cur_month_idx) if by_month[i]["cost"] > 0]
        else:
            done_months = [m["cost"] for m in by_month if m["cost"] > 0]
        if done_months:
            avg_m = sum(done_months) / len(done_months)
            forecast = total_spent + avg_m * max(0, 12 - (cur_month_idx if year == datetime.now().year else 12))
        else:
            forecast = expected_annual

        # Факт прошлого года
        prev_year = year - 1
        prev_spent = 0
        for m in range(12):
            c = db.execute("""SELECT COALESCE(SUM(r.qty*p.price),0) FROM records r
                JOIN printers p ON r.printer_id=p.id WHERE r.year=? AND r.month=? AND p.active=1""",
                (prev_year, m)).fetchone()[0]
            tc = db.execute("""SELECT COALESCE(SUM(r.qty*t.price),0) FROM toner_records r
                JOIN toners t ON r.toner_id=t.id WHERE r.year=? AND r.month=? AND t.active=1""",
                (prev_year, m)).fetchone()[0]
            prev_spent += c + tc

    avg_month = (total_spent / len(done_months)) if done_months else 0

    return jsonify({
        "total_spent": total_spent,
        "total_items": total_items,
        "expected_annual": expected_annual,
        "printer_count": len(printers),
        "by_month": by_month,
        "by_floor": by_floor,
        "by_dept": by_dept,
        "top_printers": [dict(r) for r in top_printers],
        "forecast": round(forecast),
        "prev_year_spent": prev_spent,
        "avg_month": round(avg_month),
        "months": MONTHS,
        "year": year,
    })

# ── API: Printers ─────────────────────────────────────────────────────────────
@app.route("/api/printers")
@login_required
def api_printers():
    with get_db() as db:
        rows = db.execute("SELECT * FROM printers WHERE active=1 ORDER BY floor,id").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/cartridge-names")
@login_required
def api_cartridge_names():
    with get_db() as db:
        from_stock = db.execute("SELECT cartridge FROM stock ORDER BY cartridge").fetchall()
        from_printers = db.execute("SELECT DISTINCT cartridge FROM printers WHERE active=1 ORDER BY cartridge").fetchall()
    names = sorted(set(
        [r["cartridge"] for r in from_stock] +
        [r["cartridge"] for r in from_printers]
    ))
    return jsonify(names)

def _ensure_cartridge_in_stock(db, cartridge, price):
    db.execute(
        "INSERT OR IGNORE INTO stock (cartridge,qty,min_qty,price,updated_at,updated_by) VALUES (?,0,2,?,datetime('now'),?)",
        (cartridge, price, session.get("username", ""))
    )

@app.route("/api/printers", methods=["POST"])
@admin_required
def api_add_printer():
    d = request.json
    price = int(d.get("price", 600))
    with get_db() as db:
        db.execute("INSERT INTO printers (floor,dept,model,cartridge,monthly_rate,price,note) VALUES (?,?,?,?,?,?,?)",
                   (d["floor"],d["dept"],d["model"],d["cartridge"],
                    float(d.get("monthly_rate",0.5)), price, d.get("note","")))
        _ensure_cartridge_in_stock(db, d["cartridge"], price)
    return jsonify({"ok": True})

@app.route("/api/printers/<int:pid>", methods=["PUT"])
@admin_required
def api_edit_printer(pid):
    d = request.json
    price = int(d.get("price", 600))
    with get_db() as db:
        db.execute("""UPDATE printers SET floor=?,dept=?,model=?,cartridge=?,
                      monthly_rate=?,price=?,note=? WHERE id=?""",
                   (d["floor"],d["dept"],d["model"],d["cartridge"],
                    float(d.get("monthly_rate",0.5)), price,
                    d.get("note",""), pid))
        _ensure_cartridge_in_stock(db, d["cartridge"], price)
    return jsonify({"ok": True})

@app.route("/api/printers/<int:pid>", methods=["DELETE"])
@admin_required
def api_delete_printer(pid):
    with get_db() as db:
        db.execute("UPDATE printers SET active=0 WHERE id=?", (pid,))
    return jsonify({"ok": True})

# ── API: Records ──────────────────────────────────────────────────────────────
@app.route("/api/records")
@login_required
def api_records():
    year = request.args.get("year", datetime.now().year, type=int)
    floor = request.args.get("floor", "")
    with get_db() as db:
        q = "SELECT * FROM printers WHERE active=1"
        params = []
        if floor:
            q += " AND floor=?"
            params.append(floor)
        q += " ORDER BY floor, id"
        printers = db.execute(q, params).fetchall()
        recs = db.execute("SELECT * FROM records WHERE year=?", (year,)).fetchall()
    rec_map = {(r["printer_id"], r["month"]): r["qty"] for r in recs}
    result = []
    for p in printers:
        monthly = {m: rec_map.get((p["id"], m), 0) for m in range(12)}
        total_qty = sum(monthly.values())
        result.append({**dict(p), "monthly": monthly,
                       "total_qty": total_qty, "total_cost": total_qty * p["price"]})
    return jsonify(result)

@app.route("/api/records", methods=["POST"])
@login_required
def api_set_record():
    if session.get("role") == "viewer":
        return jsonify({"error": "Нет прав на изменение"}), 403
    d = request.json
    with get_db() as db:
        db.execute("""INSERT INTO records (printer_id,year,month,qty,updated_by)
                      VALUES (?,?,?,?,?)
                      ON CONFLICT(printer_id,year,month) DO UPDATE SET qty=excluded.qty, updated_by=excluded.updated_by""",
                   (d["printer_id"], d["year"], d["month"], d["qty"], session["username"]))
    return jsonify({"ok": True})

# ── API: Toners ───────────────────────────────────────────────────────────────
@app.route("/api/toners")
@login_required
def api_toners():
    year = request.args.get("year", datetime.now().year, type=int)
    with get_db() as db:
        toners = db.execute("SELECT * FROM toners WHERE active=1 ORDER BY id").fetchall()
        recs = db.execute("SELECT * FROM toner_records WHERE year=?", (year,)).fetchall()
    rec_map = {(r["toner_id"], r["month"]): r["qty"] for r in recs}
    result = []
    for t in toners:
        monthly = {m: rec_map.get((t["id"], m), 0) for m in range(12)}
        total_qty = sum(monthly.values())
        result.append({**dict(t), "monthly": monthly,
                       "total_qty": total_qty, "total_cost": total_qty * t["price"]})
    return jsonify(result)

@app.route("/api/toners", methods=["POST"])
@admin_required
def api_add_toner():
    d = request.json
    with get_db() as db:
        db.execute("INSERT INTO toners (name,monthly_rate,price) VALUES (?,?,?)",
                   (d["name"], float(d.get("monthly_rate",0.3)), int(d.get("price",1300))))
    return jsonify({"ok": True})

@app.route("/api/toners/<int:tid>", methods=["PUT"])
@admin_required
def api_edit_toner(tid):
    d = request.json
    with get_db() as db:
        db.execute("UPDATE toners SET name=?,monthly_rate=?,price=? WHERE id=?",
                   (d["name"], float(d.get("monthly_rate",0.3)), int(d.get("price",1300)), tid))
    return jsonify({"ok": True})

@app.route("/api/toners/<int:tid>", methods=["DELETE"])
@admin_required
def api_delete_toner(tid):
    with get_db() as db:
        db.execute("UPDATE toners SET active=0 WHERE id=?", (tid,))
    return jsonify({"ok": True})

@app.route("/api/toner_records", methods=["POST"])
@login_required
def api_set_toner_record():
    if session.get("role") == "viewer":
        return jsonify({"error": "Нет прав"}), 403
    d = request.json
    with get_db() as db:
        db.execute("""INSERT INTO toner_records (toner_id,year,month,qty,updated_by)
                      VALUES (?,?,?,?,?)
                      ON CONFLICT(toner_id,year,month) DO UPDATE SET qty=excluded.qty, updated_by=excluded.updated_by""",
                   (d["toner_id"], d["year"], d["month"], d["qty"], session["username"]))
    return jsonify({"ok": True})

# ── API: Analytics ────────────────────────────────────────────────────────────
@app.route("/api/analytics")
@login_required
def api_analytics():
    year = request.args.get("year", datetime.now().year, type=int)
    with get_db() as db:
        # By cartridge
        by_cart = db.execute("""SELECT p.cartridge, SUM(r.qty) as qty, SUM(r.qty*p.price) as cost
            FROM records r JOIN printers p ON r.printer_id=p.id
            WHERE r.year=? AND p.active=1 GROUP BY p.cartridge ORDER BY cost DESC""", (year,)).fetchall()
        # By dept
        by_dept = db.execute("""SELECT p.dept, SUM(r.qty) as qty, SUM(r.qty*p.price) as cost
            FROM records r JOIN printers p ON r.printer_id=p.id
            WHERE r.year=? AND p.active=1 GROUP BY p.dept ORDER BY cost DESC LIMIT 15""", (year,)).fetchall()
        # Monthly plan vs fact
        printers = db.execute("SELECT * FROM printers WHERE active=1").fetchall()
        toners = db.execute("SELECT * FROM toners WHERE active=1").fetchall()
        monthly_plan = sum(p["monthly_rate"]*p["price"] for p in printers) + \
                       sum(t["monthly_rate"]*t["price"] for t in toners)
        monthly_fact = []
        for m in range(12):
            c = db.execute("""SELECT COALESCE(SUM(r.qty*p.price),0) FROM records r
                JOIN printers p ON r.printer_id=p.id WHERE r.year=? AND r.month=? AND p.active=1""", (year,m)).fetchone()[0]
            tc = db.execute("""SELECT COALESCE(SUM(r.qty*t.price),0) FROM toner_records r
                JOIN toners t ON r.toner_id=t.id WHERE r.year=? AND r.month=? AND t.active=1""", (year,m)).fetchone()[0]
            monthly_fact.append(c+tc)

    return jsonify({
        "by_cartridge": [dict(r) for r in by_cart],
        "by_dept": [dict(r) for r in by_dept],
        "monthly_plan": monthly_plan,
        "monthly_fact": monthly_fact,
        "months": MONTHS,
    })

# ── API: Users (admin) ────────────────────────────────────────────────────────
@app.route("/api/users")
@admin_required
def api_users():
    with get_db() as db:
        rows = db.execute("SELECT id,username,role,created_at FROM users ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users", methods=["POST"])
@admin_required
def api_add_user():
    d = request.json
    try:
        with get_db() as db:
            db.execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                       (d["username"], hash_pw(d["password"]), d.get("role","viewer")))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/users/<int:uid>", methods=["PUT"])
@admin_required
def api_edit_user(uid):
    d = request.json
    with get_db() as db:
        if d.get("password"):
            db.execute("UPDATE users SET role=?,password=? WHERE id=?",
                       (d["role"], hash_pw(d["password"]), uid))
        else:
            db.execute("UPDATE users SET role=? WHERE id=?", (d["role"], uid))
    return jsonify({"ok": True})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@admin_required
def api_delete_user(uid):
    if uid == session["user_id"]:
        return jsonify({"error": "Нельзя удалить себя"}), 400
    with get_db() as db:
        db.execute("DELETE FROM users WHERE id=?", (uid,))
    return jsonify({"ok": True})

# ── API: Backup ───────────────────────────────────────────────────────────────
@app.route("/api/backup", methods=["POST"])
@admin_required
def api_backup():
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(BACKUP_DIR, f"data_{ts}.db")
    shutil.copy2(DB_PATH, dest)
    # Keep only last 30 backups
    files = sorted(os.listdir(BACKUP_DIR))
    for f in files[:-30]:
        os.remove(os.path.join(BACKUP_DIR, f))
    return jsonify({"ok": True, "file": f"data_{ts}.db"})

@app.route("/api/backup/download/<filename>")
@admin_required
def api_backup_download(filename):
    path = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({"error": "Файл не найден"}), 404

@app.route("/api/backups")
@admin_required
def api_backups():
    files = sorted(os.listdir(BACKUP_DIR), reverse=True)
    result = []
    for f in files[:20]:
        p = os.path.join(BACKUP_DIR, f)
        result.append({"name": f, "size": os.path.getsize(p),
                       "modified": datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d.%m.%Y %H:%M")})
    return jsonify(result)


# -- API: Stock
@app.route("/api/stock")
@login_required
def api_stock():
    with get_db() as db:
        all_carts = db.execute(
            "SELECT DISTINCT cartridge FROM printers WHERE active=1 ORDER BY cartridge"
        ).fetchall()
        stock_rows = db.execute("SELECT * FROM stock").fetchall()
        stock_map = {r["cartridge"]: dict(r) for r in stock_rows}
        result = []
        for row in all_carts:
            cart = row["cartridge"]
            printers = db.execute(
                "SELECT id, floor, dept, model FROM printers WHERE cartridge=? AND active=1 ORDER BY floor, id",
                (cart,)
            ).fetchall()
            s = stock_map.get(cart, {"qty": 0, "min_qty": 2, "price": 0, "note": "", "updated_at": None, "updated_by": ""})
            result.append({
                "cartridge": cart,
                "qty": s["qty"],
                "min_qty": s["min_qty"],
                "price": s["price"],
                "note": s["note"],
                "updated_at": s["updated_at"],
                "updated_by": s["updated_by"],
                "low": s["qty"] <= s["min_qty"],
                "empty": s["qty"] == 0,
                "printers": [dict(p) for p in printers],
                "printer_count": len(printers),
            })
    return jsonify(result)

@app.route("/api/stock", methods=["POST"])
@login_required
def api_update_stock():
    if session.get("role") == "viewer":
        return jsonify({"error": "No permission"}), 403
    d = request.json
    cart = d["cartridge"]
    qty = int(d.get("qty", 0))
    min_qty = int(d.get("min_qty", 2))
    price = int(d.get("price", 0))
    note = d.get("note", "")
    with get_db() as db:
        db.execute(
            "INSERT INTO stock (cartridge,qty,min_qty,price,note,updated_at,updated_by) VALUES (?,?,?,?,?,datetime('now'),?)"
            " ON CONFLICT(cartridge) DO UPDATE SET qty=excluded.qty,min_qty=excluded.min_qty,"
            "price=excluded.price,note=excluded.note,updated_at=excluded.updated_at,updated_by=excluded.updated_by",
            (cart, qty, min_qty, price, note, session["username"]))
    return jsonify({"ok": True})

@app.route("/api/stock/move", methods=["POST"])
@login_required
def api_stock_move():
    if session.get("role") == "viewer":
        return jsonify({"error": "No permission"}), 403
    d = request.json
    cart = d["cartridge"]
    change = int(d["change"])
    reason = d.get("reason", "")
    with get_db() as db:
        existing = db.execute("SELECT qty FROM stock WHERE cartridge=?", (cart,)).fetchone()
        if existing is None:
            db.execute("INSERT INTO stock (cartridge,qty,updated_at,updated_by) VALUES (?,?,datetime('now'),?)",
                       (cart, max(0, change), session["username"]))
        else:
            new_qty = max(0, existing["qty"] + change)
            db.execute("UPDATE stock SET qty=?,updated_at=datetime('now'),updated_by=? WHERE cartridge=?",
                       (new_qty, session["username"], cart))
        db.execute("INSERT INTO stock_log (cartridge,change_qty,reason,created_by) VALUES (?,?,?,?)",
                   (cart, change, reason, session["username"]))
    return jsonify({"ok": True})

@app.route("/api/stock/log")
@login_required
def api_stock_log():
    cart = request.args.get("cartridge", "")
    with get_db() as db:
        if cart:
            rows = db.execute("SELECT * FROM stock_log WHERE cartridge=? ORDER BY id DESC LIMIT 50", (cart,)).fetchall()
        else:
            rows = db.execute("SELECT * FROM stock_log ORDER BY id DESC LIMIT 100").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/stock/alerts")
@login_required
def api_stock_alerts():
    with get_db() as db:
        low = db.execute("SELECT cartridge,qty,min_qty FROM stock WHERE qty <= min_qty AND qty > 0").fetchall()
        empty = db.execute("SELECT cartridge,qty,min_qty FROM stock WHERE qty = 0").fetchall()
    return jsonify({"low": [dict(r) for r in low], "empty": [dict(r) for r in empty], "count": len(low)+len(empty)})


# -- API: Equipment
@app.route("/api/equipment")
@login_required
def api_equipment():
    floor = request.args.get("floor", "")
    dept  = request.args.get("dept", "")
    status = request.args.get("status", "")
    search = request.args.get("search", "")
    with get_db() as db:
        q = "SELECT * FROM equipment WHERE 1=1"
        params = []
        if floor:  q += " AND floor=?";  params.append(floor)
        if dept:   q += " AND dept=?";   params.append(dept)
        if status: q += " AND status=?"; params.append(status)
        if search:
            q += " AND (pc_name LIKE ? OR dept LIKE ? OR responsible LIKE ? OR pc_inv LIKE ? OR pc_ip LIKE ? OR pc_brand LIKE ? OR pc_model LIKE ?)"
            s = "%" + search + "%"
            params += [s,s,s,s,s,s,s]
        q += " ORDER BY floor, dept, id"
        rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/equipment", methods=["POST"])
@admin_required
def api_add_equipment():
    d = request.json
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        db.execute("""INSERT INTO equipment
            (floor,dept,responsible,pc_name,pc_inv,pc_brand,pc_model,pc_serial,pc_os,
             pc_cpu,pc_ram,pc_hdd,pc_ip,pc_mac,monitor,monitor_inv,monitor_serial,
             keyboard,mouse,ups,phone,other_devices,purchase_date,warranty_until,
             status,note,antivirus,created_at,updated_at,updated_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get("floor",""), d.get("dept",""), d.get("responsible",""),
             d.get("pc_name",""), d.get("pc_inv",""), d.get("pc_brand",""),
             d.get("pc_model",""), d.get("pc_serial",""), d.get("pc_os",""),
             d.get("pc_cpu",""), d.get("pc_ram",""), d.get("pc_hdd",""),
             d.get("pc_ip",""), d.get("pc_mac",""), d.get("monitor",""),
             d.get("monitor_inv",""), d.get("monitor_serial",""),
             d.get("keyboard",""), d.get("mouse",""), d.get("ups",""),
             d.get("phone",""), d.get("other_devices",""),
             d.get("purchase_date",""), d.get("warranty_until",""),
             d.get("status","active"), d.get("note",""), d.get("antivirus",""),
             now, now, session["username"]))
    return jsonify({"ok": True})

@app.route("/api/equipment/<int:eid>", methods=["PUT"])
@admin_required
def api_edit_equipment(eid):
    d = request.json
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        old = db.execute("SELECT * FROM equipment WHERE id=?", (eid,)).fetchone()
        if old:
            fields = ["floor","dept","responsible","pc_name","pc_inv","pc_brand","pc_model",
                      "pc_serial","pc_os","pc_cpu","pc_ram","pc_hdd","pc_ip","pc_mac",
                      "monitor","monitor_inv","monitor_serial","keyboard","mouse","ups",
                      "phone","other_devices","purchase_date","warranty_until","status","note","antivirus"]
            for f in fields:
                ov = old[f] or ""
                nv = d.get(f, "") or ""
                if str(ov) != str(nv):
                    db.execute("INSERT INTO equipment_history (equipment_id,field_name,old_value,new_value,changed_at,changed_by) VALUES (?,?,?,?,?,?)",
                               (eid, f, ov, nv, now, session["username"]))
        db.execute("""UPDATE equipment SET
            floor=?,dept=?,responsible=?,pc_name=?,pc_inv=?,pc_brand=?,pc_model=?,
            pc_serial=?,pc_os=?,pc_cpu=?,pc_ram=?,pc_hdd=?,pc_ip=?,pc_mac=?,
            monitor=?,monitor_inv=?,monitor_serial=?,keyboard=?,mouse=?,ups=?,
            phone=?,other_devices=?,purchase_date=?,warranty_until=?,
            status=?,note=?,antivirus=?,updated_at=?,updated_by=? WHERE id=?""",
            (d.get("floor",""), d.get("dept",""), d.get("responsible",""),
             d.get("pc_name",""), d.get("pc_inv",""), d.get("pc_brand",""),
             d.get("pc_model",""), d.get("pc_serial",""), d.get("pc_os",""),
             d.get("pc_cpu",""), d.get("pc_ram",""), d.get("pc_hdd",""),
             d.get("pc_ip",""), d.get("pc_mac",""), d.get("monitor",""),
             d.get("monitor_inv",""), d.get("monitor_serial",""),
             d.get("keyboard",""), d.get("mouse",""), d.get("ups",""),
             d.get("phone",""), d.get("other_devices",""),
             d.get("purchase_date",""), d.get("warranty_until",""),
             d.get("status","active"), d.get("note",""), d.get("antivirus",""),
             now, session["username"], eid))
    return jsonify({"ok": True})

@app.route("/api/equipment/<int:eid>", methods=["DELETE"])
@admin_required
def api_delete_equipment(eid):
    with get_db() as db:
        db.execute("UPDATE equipment SET status='written_off' WHERE id=?", (eid,))
    return jsonify({"ok": True})

@app.route("/api/equipment/<int:eid>/history")
@login_required
def api_equipment_history(eid):
    with get_db() as db:
        rows = db.execute("SELECT * FROM equipment_history WHERE equipment_id=? ORDER BY id DESC LIMIT 100", (eid,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/equipment/stats")
@login_required
def api_equipment_stats():
    with get_db() as db:
        total   = db.execute("SELECT COUNT(*) FROM equipment WHERE status!='written_off'").fetchone()[0]
        active  = db.execute("SELECT COUNT(*) FROM equipment WHERE status='active'").fetchone()[0]
        repair  = db.execute("SELECT COUNT(*) FROM equipment WHERE status='repair'").fetchone()[0]
        storage = db.execute("SELECT COUNT(*) FROM equipment WHERE status='storage'").fetchone()[0]
        written = db.execute("SELECT COUNT(*) FROM equipment WHERE status='written_off'").fetchone()[0]
        by_floor = db.execute("SELECT floor, COUNT(*) as cnt FROM equipment WHERE status!='written_off' GROUP BY floor ORDER BY floor").fetchall()
        # Warranty expiring within 90 days
        import datetime as dt
        soon = (dt.datetime.now() + dt.timedelta(days=90)).strftime("%Y-%m-%d")
        today = dt.datetime.now().strftime("%Y-%m-%d")
        warranty_soon = db.execute(
            "SELECT COUNT(*) FROM equipment WHERE warranty_until!='' AND warranty_until<=? AND warranty_until>=? AND status='active'",
            (soon, today)).fetchone()[0]
        no_antivirus = db.execute(
            "SELECT COUNT(*) FROM equipment WHERE status='active' AND (antivirus IS NULL OR antivirus='')").fetchone()[0]
        with_antivirus = db.execute(
            "SELECT COUNT(*) FROM equipment WHERE status='active' AND antivirus!='' AND antivirus IS NOT NULL").fetchone()[0]
    return jsonify({"total":total,"active":active,"repair":repair,"storage":storage,
                    "written_off":written,"by_floor":[dict(r) for r in by_floor],
                    "warranty_soon":warranty_soon,
                    "no_antivirus":no_antivirus,"with_antivirus":with_antivirus})

@app.route("/api/equipment/depts")
@login_required
def api_equipment_depts():
    with get_db() as db:
        rows = db.execute("SELECT DISTINCT dept FROM equipment WHERE status!='written_off' ORDER BY dept").fetchall()
    return jsonify([r["dept"] for r in rows])

@app.route("/api/equipment/win11")
@login_required
def api_equipment_win11():
    with get_db() as db:
        equip = db.execute("""SELECT id, pc_cpu, pc_ram, pc_hdd, dept, floor,
            responsible, pc_ip, pc_brand, pc_model, pc_os
            FROM equipment WHERE status!='written_off'""").fetchall()
    counts = {"compatible": 0, "needs_upgrade": 0, "incompatible": 0, "unknown": 0}
    details = []
    for eq in equip:
        status, reason = check_win11_compat(eq["pc_cpu"], eq["pc_ram"])
        counts[status] += 1
        details.append({
            "id": eq["id"],
            "dept": eq["dept"],
            "floor": eq["floor"],
            "responsible": eq["responsible"],
            "pc_ip": eq["pc_ip"],
            "pc_brand": eq["pc_brand"],
            "pc_model": eq["pc_model"],
            "pc_os": eq["pc_os"],
            "cpu": eq["pc_cpu"],
            "ram": eq["pc_ram"],
            "status": status,
            "reason": reason,
        })
    details.sort(key=lambda x: ["incompatible","needs_upgrade","unknown","compatible"].index(x["status"]))
    return jsonify({**counts, "total": len(details), "details": details})


# ── SSE: Live updates endpoint ────────────────────────────────────────────────
@app.route("/api/floors")
@login_required
def api_floors():
    with get_db() as db:
        rows = db.execute("SELECT DISTINCT floor FROM printers WHERE active=1 ORDER BY floor").fetchall()
    return jsonify([r["floor"] for r in rows])


# ── API: Export ───────────────────────────────────────────────────────────────
@app.route("/export/excel/<report>")
@login_required
def export_excel(report):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    year = request.args.get("year", datetime.now().year, type=int)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Styles
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    sub_fill  = PatternFill("solid", fgColor="2D4A7A")
    alt_fill  = PatternFill("solid", fgColor="F0F4FF")
    title_font = Font(bold=True, size=14, color="1E3A5F")
    bold_font  = Font(bold=True, size=10)
    wrap_align = Alignment(wrap_text=True, vertical="center")
    center     = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC")
    )

    def set_hdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin

    def set_cell(ws, r, c, val, bold=False, fill=None, align=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, size=10)
        cell.border = thin
        cell.alignment = align or Alignment(vertical="center")
        if fill: cell.fill = fill
        return cell

    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    if report == "records":
        ws.title = "Расход картриджей"
        with get_db() as db:
            printers = db.execute("SELECT * FROM printers WHERE active=1 ORDER BY floor,id").fetchall()
            recs = db.execute("SELECT * FROM records WHERE year=?", (year,)).fetchall()
        rec_map = {(r["printer_id"], r["month"]): r["qty"] for r in recs}

        # Title
        ws.merge_cells("A1:T1")
        t = ws["A1"]; t.value = f"РАСХОД КАРТРИДЖЕЙ И ТОНЕРОВ — {year} год"; t.font = title_font; t.alignment = center
        ws.merge_cells("A2:T2")
        ws["A2"].value = f"Дата выгрузки: {now_str}"; ws["A2"].font = Font(italic=True, size=9, color="888888"); ws["A2"].alignment = center
        ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 16

        MONTHS_SHORT = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
        headers = ["№","Этаж","Помещение","Модель принтера","Картридж","Цена,руб"] + MONTHS_SHORT + ["Итого шт.","Сумма руб."]
        set_hdr(ws, 3, headers)
        ws.row_dimensions[3].height = 30

        row = 4; prev_floor = None
        total_qty = 0; total_cost = 0
        for i, p in enumerate(printers):
            if p["floor"] != prev_floor:
                prev_floor = p["floor"]
                ws.merge_cells(f"A{row}:T{row}")
                fc = ws.cell(row=row, column=1, value=f"  {p['floor']}")
                fc.font = Font(bold=True, size=10, color="FFFFFF")
                fc.fill = PatternFill("solid", fgColor="344054")
                fc.alignment = Alignment(vertical="center")
                for c in range(1,21): ws.cell(row=row,column=c).fill = PatternFill("solid",fgColor="344054"); ws.cell(row=row,column=c).border=thin
                ws.row_dimensions[row].height = 20; row += 1

            mq = {m: rec_map.get((p["id"], m), 0) for m in range(12)}
            qty = sum(mq.values()); cost = qty * p["price"]
            total_qty += qty; total_cost += cost
            fill = alt_fill if i % 2 == 0 else None
            set_cell(ws,row,1,i+1,align=center,fill=fill)
            set_cell(ws,row,2,p["floor"],fill=fill)
            set_cell(ws,row,3,p["dept"],fill=fill)
            set_cell(ws,row,4,p["model"],fill=fill)
            set_cell(ws,row,5,p["cartridge"],fill=fill)
            set_cell(ws,row,6,p["price"],align=center,fill=fill)
            for mi in range(12):
                c = ws.cell(row=row, column=7+mi, value=mq[mi] if mq[mi] else None)
                c.alignment = center; c.border = thin; c.fill = fill or PatternFill()
                if mq[mi]: c.font = Font(bold=True, color="1A6B3A", size=10)
            set_cell(ws,row,19,qty if qty else None,bold=bool(qty),align=center,fill=fill)
            c = ws.cell(row=row,column=20,value=cost if cost else None)
            c.number_format = '#,##0 "руб."'; c.alignment=center; c.border=thin; c.fill=fill or PatternFill()
            if cost: c.font = Font(bold=True, color="1E3A5F", size=10)
            ws.row_dimensions[row].height = 18; row += 1

        # Totals
        for c in range(1,21):
            ws.cell(row=row,column=c).fill=PatternFill("solid",fgColor="1E3A5F"); ws.cell(row=row,column=c).border=thin
        t1=ws.cell(row=row,column=1,value="ИТОГО:"); t1.font=Font(bold=True,color="FFFFFF",size=10); t1.alignment=center
        ws.merge_cells(f"A{row}:R{row}")
        t2=ws.cell(row=row,column=19,value=total_qty); t2.font=Font(bold=True,color="FFFFFF",size=10); t2.alignment=center
        t3=ws.cell(row=row,column=20,value=total_cost); t3.font=Font(bold=True,color="FFFFFF",size=10); t3.alignment=center; t3.number_format='#,##0 "руб."'
        ws.row_dimensions[row].height=22

        # Column widths
        widths = [4,10,26,20,12,9]+[5]*12+[8,12]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes = "A4"

    elif report == "equipment":
        ws.title = "Учёт техники"
        with get_db() as db:
            rows = db.execute("SELECT * FROM equipment WHERE status!='written_off' ORDER BY floor,dept,id").fetchall()

        ws.merge_cells("A1:W1")
        t=ws["A1"]; t.value="РЕЕСТР КОМПЬЮТЕРНОЙ ТЕХНИКИ"; t.font=title_font; t.alignment=center
        ws.merge_cells("A2:W2")
        ws["A2"].value=f"Дата выгрузки: {now_str}  |  Всего: {len(rows)} ед."; ws["A2"].font=Font(italic=True,size=9,color="888888"); ws["A2"].alignment=center
        ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=16

        headers=["№","Этаж","Отдел","Ответственный","Инв. №","Производитель","Модель","Серийный №","ОС","Антивирус","CPU","RAM","Диск","IP-адрес","Монитор","Инв.монитора","ИБП","Телефон","Прочее","Статус","Гарантия до","Дата покупки","Примечание"]
        set_hdr(ws,3,headers); ws.row_dimensions[3].height=30

        STATUS_RU={"active":"Активно","repair":"В ремонте","storage":"На хранении","written_off":"Списано"}
        for i,r in enumerate(rows):
            fill=alt_fill if i%2==0 else None
            vals=[i+1,r["floor"],r["dept"],r["responsible"],r["pc_inv"],r["pc_brand"],r["pc_model"],
                  r["pc_serial"],r["pc_os"],r["antivirus"] or "",r["pc_cpu"],r["pc_ram"],r["pc_hdd"],r["pc_ip"],
                  r["monitor"],r["monitor_inv"],r["ups"],r["phone"],r["other_devices"],
                  STATUS_RU.get(r["status"],r["status"]),r["warranty_until"],r["purchase_date"],r["note"]]
            for c,v in enumerate(vals,1):
                set_cell(ws,4+i,c,v,fill=fill,align=wrap_align)

        widths=[4,9,22,18,12,14,16,14,16,14,16,10,12,14,18,14,14,12,20,12,12,12,22]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A4"

    elif report == "stock":
        ws.title = "Склад картриджей"
        with get_db() as db:
            all_carts=db.execute("SELECT DISTINCT cartridge FROM printers WHERE active=1 ORDER BY cartridge").fetchall()
            stock_rows=db.execute("SELECT * FROM stock").fetchall()
        sm={r["cartridge"]:dict(r) for r in stock_rows}

        ws.merge_cells("A1:F1")
        t=ws["A1"]; t.value="ОСТАТКИ КАРТРИДЖЕЙ НА СКЛАДЕ"; t.font=title_font; t.alignment=center
        ws.merge_cells("A2:F2")
        ws["A2"].value=f"Дата выгрузки: {now_str}"; ws["A2"].font=Font(italic=True,size=9,color="888888"); ws["A2"].alignment=center

        set_hdr(ws,3,["Картридж","Остаток (шт.)","Мин. запас","Статус","Цена, руб.","Примечание"])
        for i,row in enumerate(all_carts):
            c=row["cartridge"]; s=sm.get(c,{"qty":0,"min_qty":2,"price":0,"note":""})
            qty=s["qty"]; mn=s["min_qty"]
            status="Нет в наличии" if qty==0 else ("Заканчивается" if qty<=mn else "В норме")
            fill=PatternFill("solid",fgColor="FFE5E5") if qty==0 else (PatternFill("solid",fgColor="FFF8E5") if qty<=mn else (alt_fill if i%2==0 else None))
            set_cell(ws,4+i,1,c,fill=fill)
            c2=ws.cell(row=4+i,column=2,value=qty); c2.alignment=center; c2.border=thin; c2.fill=fill or PatternFill()
            c2.font=Font(bold=True,size=10,color=("C0392B" if qty==0 else ("E67E22" if qty<=mn else "1A6B3A")))
            set_cell(ws,4+i,3,mn,align=center,fill=fill)
            set_cell(ws,4+i,4,status,fill=fill)
            set_cell(ws,4+i,5,s["price"] if s["price"] else None,fill=fill)
            set_cell(ws,4+i,6,s.get("note",""),fill=fill)
            ws.row_dimensions[4+i].height=18
        for i,w in enumerate([20,14,12,16,12,28],1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A4"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"{report}_{year}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/export/print/<report>")
@login_required
def export_print(report):
    year = request.args.get("year", datetime.now().year, type=int)
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    MONTHS_SHORT = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
    STATUS_RU = {"active":"Активно","repair":"В ремонте","storage":"На хранении","written_off":"Списано"}

    css = """
    <style>
      *{box-sizing:border-box;margin:0;padding:0}
      body{font-family:Arial,sans-serif;font-size:10px;color:#000;background:#fff;padding:15mm}
      h1{font-size:16px;text-align:center;margin-bottom:4px;color:#1E3A5F}
      .sub{text-align:center;color:#888;font-size:9px;margin-bottom:14px}
      table{width:100%;border-collapse:collapse;margin-bottom:20px}
      th{background:#1E3A5F;color:#fff;padding:5px 6px;text-align:left;font-size:9px;border:1px solid #2D4A7A}
      td{padding:4px 6px;border:1px solid #ccc;vertical-align:middle;font-size:9px;word-wrap:break-word;overflow-wrap:break-word;max-width:200px}
      tr:nth-child(even) td{background:#F0F4FF}
      .floor-row td{background:#344054;color:#fff;font-weight:bold;font-size:9px}
      .total-row td{background:#1E3A5F;color:#fff;font-weight:bold}
      .qty{font-weight:bold;color:#1A6B3A;text-align:center}
      .cost{font-weight:bold;color:#1E3A5F;text-align:right}
      .center{text-align:center}
      .red{color:#C0392B;font-weight:bold}
      .orange{color:#E67E22;font-weight:bold}
      .green{color:#1A6B3A;font-weight:bold}
      .badge{display:inline-block;padding:1px 6px;border-radius:10px;font-size:8px;font-weight:bold}
      .badge-green{background:#d4edda;color:#155724}
      .badge-red{background:#f8d7da;color:#721c24}
      .badge-orange{background:#fff3cd;color:#856404}
      .badge-gray{background:#e2e3e5;color:#383d41}
      @media print{
        body{padding:10mm}
        .no-print{display:none}
        @page{size:A4 landscape;margin:10mm}
      }
      .print-btn{position:fixed;top:15px;right:15px;padding:8px 16px;background:#1E3A5F;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px;z-index:100}
    </style>
    <button class="print-btn no-print" onclick="window.print()">🖨️ Печать / PDF</button>
    """

    if report == "records":
        with get_db() as db:
            printers = db.execute("SELECT * FROM printers WHERE active=1 ORDER BY floor,id").fetchall()
            recs = db.execute("SELECT * FROM records WHERE year=?", (year,)).fetchall()
        rec_map = {(r["printer_id"], r["month"]): r["qty"] for r in recs}
        total_qty=0; total_cost=0

        html = css + f"<h1>РАСХОД КАРТРИДЖЕЙ И ТОНЕРОВ — {year} год</h1><div class='sub'>Дата выгрузки: {now_str}</div>"
        html += "<table><thead><tr><th>№</th><th>Этаж</th><th>Помещение</th><th>Модель</th><th>Картридж</th><th>Цена</th>"
        html += "".join(f"<th class='center'>{m}</th>" for m in MONTHS_SHORT)
        html += "<th class='center'>Итого</th><th class='center'>Сумма</th></tr></thead><tbody>"

        prev_floor=None; idx=0
        for p in printers:
            if p["floor"]!=prev_floor:
                prev_floor=p["floor"]
                html+=f"<tr class='floor-row'><td colspan='20'>{p['floor']}</td></tr>"
            idx+=1
            mq={m:rec_map.get((p["id"],m),0) for m in range(12)}
            qty=sum(mq.values()); cost=qty*p["price"]
            total_qty+=qty; total_cost+=cost
            html+=f"<tr><td class='center'>{idx}</td><td>{p['floor']}</td><td>{p['dept']}</td><td>{p['model']}</td><td><b>{p['cartridge']}</b></td><td class='center'>{p['price']}</td>"
            for mi in range(12):
                html+=f"<td class='center{' qty' if mq[mi] else ''}'>{mq[mi] if mq[mi] else ''}</td>"
            html+=f"<td class='center qty'>{qty if qty else ''}</td><td class='cost'>{f'{total_cost:,}'.replace(',', ' ')} ₽" if cost else "</td><td></td>"
            html+="</tr>"
        html+=f"<tr class='total-row'><td colspan='18' style='text-align:right'>ИТОГО:</td><td class='center'>{total_qty}</td><td class='center'>{f'{total_cost:,}'.replace(',', ' ')} ₽</td></tr>"
        html+="</tbody></table>"

    elif report == "equipment":
        with get_db() as db:
            rows=db.execute("SELECT * FROM equipment ORDER BY floor,dept,id").fetchall()
        html = css + f"<h1>РЕЕСТР КОМПЬЮТЕРНОЙ ТЕХНИКИ</h1><div class='sub'>Дата выгрузки: {now_str} | Всего: {len(rows)} ед.</div>"
        html += "<table><thead><tr><th>№</th><th>Этаж</th><th>Отдел</th><th>Ответственный</th><th>Инв.№</th><th>Производитель</th><th>Модель</th><th>ОС</th><th>Антивирус</th><th>CPU</th><th>RAM</th><th>IP</th><th>Монитор</th><th>ИБП</th><th>Телефон</th><th>Статус</th><th>Гарантия до</th><th>Примечание</th></tr></thead><tbody>"
        BADGE={"active":"badge-green","repair":"badge-red","storage":"badge-orange","written_off":"badge-gray"}
        for i,r in enumerate(rows):
            s=STATUS_RU.get(r["status"],r["status"]); b=BADGE.get(r["status"],"badge-gray")
            av=r["antivirus"] or ""
            av_cell=f"<span class='badge badge-green'>{av}</span>" if av else "<span class='badge badge-red'>—</span>"
            html+=f"<tr><td class='center'>{i+1}</td><td>{r['floor']}</td><td>{r['dept']}</td><td>{r['responsible']or''}</td><td>{r['pc_inv']or''}</td><td>{r['pc_brand']or''}</td><td>{r['pc_model']or''}</td><td>{r['pc_os']or''}</td><td>{av_cell}</td><td>{r['pc_cpu']or''}</td><td>{r['pc_ram']or''}</td><td>{r['pc_ip']or''}</td><td>{r['monitor']or''}</td><td>{r['ups']or''}</td><td>{r['phone']or''}</td><td><span class='badge {b}'>{s}</span></td><td>{r['warranty_until']or''}</td><td>{r['note']or''}</td></tr>"
        html+="</tbody></table>"

    elif report == "stock":
        with get_db() as db:
            carts=db.execute("SELECT DISTINCT cartridge FROM printers WHERE active=1 ORDER BY cartridge").fetchall()
            stock_rows=db.execute("SELECT * FROM stock").fetchall()
        sm={r["cartridge"]:dict(r) for r in stock_rows}
        html = css + f"<h1>ОСТАТКИ КАРТРИДЖЕЙ НА СКЛАДЕ</h1><div class='sub'>Дата выгрузки: {now_str}</div>"
        html += "<table><thead><tr><th>Картридж</th><th class='center'>Остаток (шт.)</th><th class='center'>Мин. запас</th><th>Статус</th><th class='center'>Цена, руб.</th><th>Примечание</th></tr></thead><tbody>"
        for row in carts:
            c=row["cartridge"]; s=sm.get(c,{"qty":0,"min_qty":2,"price":0,"note":""})
            qty=s["qty"]; mn=s["min_qty"]
            cls="red" if qty==0 else ("orange" if qty<=mn else "green")
            status="Нет в наличии" if qty==0 else ("Заканчивается" if qty<=mn else "В норме")
            html+=f"<tr><td><b>{c}</b></td><td class='center {cls}'>{qty}</td><td class='center'>{mn}</td><td>{status}</td><td class='center'>{s['price'] or ''}</td><td>{s.get('note','')}</td></tr>"
        html+="</tbody></table>"
    else:
        return "Unknown report", 404

    return make_response(html)

if __name__ == "__main__":
    init_db()
    print("\n" + "="*55)
    print("  CartridgeApp started!")
    print("  Open browser: http://localhost:5000")
    print("  Login: admin | Password: admin123")
    print("="*55 + "\n")
    try:
        from waitress import serve
        print("  Mode: production (waitress)")
        serve(app, host="0.0.0.0", port=5000, threads=4)
    except ImportError:
        print("  Mode: Flask dev server")
        print("  Install waitress: pip install waitress")
        app.run(host="0.0.0.0", port=5000, debug=False)
